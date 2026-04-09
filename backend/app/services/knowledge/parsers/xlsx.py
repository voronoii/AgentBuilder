from __future__ import annotations

import asyncio
from pathlib import Path

from openpyxl import load_workbook

from app.services.knowledge.parsers.base import ParsedDocument


class XlsxParser:
    async def parse(self, path: Path) -> ParsedDocument:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, self._parse_sync, path)

    def _parse_sync(self, path: Path) -> ParsedDocument:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        chunks: list[str] = []
        for sheet in wb.worksheets:
            chunks.append(f"# sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    chunks.append(" | ".join(cells))
        return ParsedDocument(
            text="\n".join(chunks),
            metadata={"sheet_count": len(wb.worksheets), "filename": path.name},
        )
